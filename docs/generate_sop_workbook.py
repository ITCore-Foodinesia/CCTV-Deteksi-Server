"""
Script untuk membuat SOP PR Review Workbook dengan multiple sheets
Jalankan: python generate_sop_workbook.py
Output: docs/SOP_PR_Review_Workbook.xlsx
"""

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call(['pip', 'install', 'openpyxl'])
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

# Styles
header_font = Font(bold=True, size=12, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
title_font = Font(bold=True, size=14)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def apply_header_style(ws, row_num, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

def auto_width(ws):
    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

# Create workbook
wb = openpyxl.Workbook()

# ============ SHEET 1: Overview Workflow ============
ws1 = wb.active
ws1.title = "1. Overview Workflow"

ws1['A1'] = "SOP PULL REQUEST REVIEW - OVERVIEW WORKFLOW"
ws1['A1'].font = title_font
ws1['A2'] = "Versi: 1.0 | Tanggal: 2026-02-01 | Owner: [Nama Maintainer]"

headers1 = ["No", "Langkah", "Deskripsi", "Lokasi/Tool", "Output", "PIC", "Catatan"]
for col, header in enumerate(headers1, 1):
    ws1.cell(row=4, column=col, value=header)
apply_header_style(ws1, 4, len(headers1))

data1 = [
    [1, "Terima Notifikasi PR", "Developer mengajukan PR dari feature branch ke main", "GitHub > Pull Requests tab", "Notifikasi email/GitHub", "Maintainer", "Cek inbox secara berkala"],
    [2, "Buka & Baca PR", "Baca judul, deskripsi, dan konteks perubahan", "GitHub > PR Detail", "Pemahaman scope perubahan", "Maintainer", "Pastikan deskripsi jelas"],
    [3, "Review Code (Diff)", "Periksa perubahan baris per baris", "GitHub > Files Changed tab", "Identifikasi issue/improvement", "Maintainer", "Gunakan checklist review"],
    [4, "Test Lokal (Opsional)", "Jalankan code di local environment", "Terminal / IDE", "Konfirmasi code berjalan", "Maintainer", "Wajib untuk perubahan kompleks"],
    [5, "Cek CI/CD Status", "Pastikan automated tests passed", "GitHub > PR Checks", "Status passed/failed", "System", "Jangan merge jika failed"],
    [6, "Berikan Feedback", "Approve, Request Changes, atau Comment", "GitHub > Review changes", "Review submitted", "Maintainer", "Pilih sesuai kondisi"],
    [7, "Merge PR", "Gabungkan code ke main branch", "GitHub > Merge button", "Code masuk ke main", "Maintainer", "Pilih merge strategy"],
    [8, "Cleanup", "Hapus branch & pull ke local", "GitHub & Terminal", "Branch bersih", "Maintainer", "Opsional tapi direkomendasikan"],
]

for row_idx, row_data in enumerate(data1, 5):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws1.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border

auto_width(ws1)

# ============ SHEET 2: Checklist Review ============
ws2 = wb.create_sheet("2. Checklist Review")

ws2['A1'] = "CHECKLIST REVIEW PR"
ws2['A1'].font = title_font
ws2['A2'] = "Instruksi: Isi kolom Status dengan ✓ (OK) atau ✗ (Perlu Fix)"

headers2 = ["No", "Item Checklist", "Status (✓/✗)", "Komentar", "Prioritas", "Kategori"]
for col, header in enumerate(headers2, 1):
    ws2.cell(row=4, column=col, value=header)
apply_header_style(ws2, 4, len(headers2))

data2 = [
    [1, "Deskripsi PR jelas dan lengkap", "", "", "Tinggi", "Dokumentasi"],
    [2, "Judul PR menjelaskan perubahan", "", "", "Tinggi", "Dokumentasi"],
    [3, "Code dapat dibaca dan dipahami", "", "", "Tinggi", "Readability"],
    [4, "Tidak ada hardcoded secrets/password", "", "", "Kritis", "Security"],
    [5, "Tidak ada console.log/print debug", "", "", "Sedang", "Clean Code"],
    [6, "Error handling sudah ada", "", "", "Tinggi", "Reliability"],
    [7, "Tidak ada bug yang terlihat", "", "", "Tinggi", "Correctness"],
    [8, "Logic sesuai dengan requirement", "", "", "Tinggi", "Correctness"],
    [9, "Tidak ada duplicate code", "", "", "Sedang", "Maintainability"],
    [10, "Naming variable/function jelas", "", "", "Sedang", "Readability"],
    [11, "Unit test ada (jika logic baru)", "", "", "Tinggi", "Testing"],
    [12, "CI/CD checks passed", "", "", "Tinggi", "Automation"],
    [13, "Tidak ada conflict dengan main", "", "", "Tinggi", "Integration"],
    [14, "Performance tidak menurun", "", "", "Sedang", "Performance"],
    [15, "Compatible dengan browser/device target", "", "", "Sedang", "Compatibility"],
]

for row_idx, row_data in enumerate(data2, 5):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border

ws2['A21'] = "RINGKASAN REVIEW"
ws2['A21'].font = Font(bold=True)
ws2['A22'] = "Total Item OK:"
ws2['A23'] = "Total Item Perlu Fix:"
ws2['A24'] = "Keputusan:"
ws2['A25'] = "Catatan Reviewer:"

auto_width(ws2)

# ============ SHEET 3: Log Review PR ============
ws3 = wb.create_sheet("3. Log Review PR")

ws3['A1'] = "LOG REVIEW PR"
ws3['A1'].font = title_font
ws3['A2'] = "Instruksi: Tambahkan baris baru setiap kali ada PR yang di-review"

headers3 = ["Tanggal", "PR Number", "PR Title", "Author", "Reviewer", "Status Review", "Merge Date", "Merge Strategy", "Catatan"]
for col, header in enumerate(headers3, 1):
    ws3.cell(row=4, column=col, value=header)
apply_header_style(ws3, 4, len(headers3))

# Sample row
sample = ["2026-02-01", "#1", "Contoh: Fix login bug", "Developer A", "Maintainer", "Approved", "2026-02-01", "Squash", "Contoh entry"]
for col_idx, value in enumerate(sample, 1):
    cell = ws3.cell(row=5, column=col_idx, value=value)
    cell.border = thin_border

# Empty rows for future entries
for row in range(6, 25):
    for col in range(1, 10):
        ws3.cell(row=row, column=col).border = thin_border

auto_width(ws3)

# ============ SHEET 4: Merge Strategy Guide ============
ws4 = wb.create_sheet("4. Merge Strategy")

ws4['A1'] = "MERGE STRATEGY GUIDE"
ws4['A1'].font = title_font

headers4 = ["Strategy", "Deskripsi", "Kapan Digunakan", "Kelebihan", "Kekurangan", "Rekomendasi"]
for col, header in enumerate(headers4, 1):
    ws4.cell(row=3, column=col, value=header)
apply_header_style(ws4, 3, len(headers4))

data4 = [
    ["Merge Commit", "Semua commit tetap ada + 1 merge commit baru", "Default untuk kebanyakan kasus", "History lengkap", "History bisa berantakan", "Team besar"],
    ["Squash and Merge", "Semua commit dijadikan 1 commit", "Merapikan history", "History bersih dan linear", "Detail commit hilang", "REKOMENDASI PEMULA"],
    ["Rebase and Merge", "Commit ditaruh di atas main", "Linear history ketat", "History paling bersih", "Mengubah commit hash", "Team yang paham Git"],
]

for row_idx, row_data in enumerate(data4, 4):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws4.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border

auto_width(ws4)

# ============ SHEET 5: Command Reference ============
ws5 = wb.create_sheet("5. Command Reference")

ws5['A1'] = "GIT COMMAND REFERENCE"
ws5['A1'].font = title_font

headers5 = ["Kategori", "Aksi", "Command", "Deskripsi"]
for col, header in enumerate(headers5, 1):
    ws5.cell(row=3, column=col, value=header)
apply_header_style(ws5, 3, len(headers5))

data5 = [
    ["Fetch", "Ambil info branch terbaru", "git fetch origin", "Mengambil semua info branch dari remote"],
    ["Checkout", "Pindah ke branch PR", "git checkout nama-branch", "Berpindah ke branch yang akan di-test"],
    ["Pull", "Update local dengan remote", "git pull origin main", "Mengambil perubahan terbaru dari remote"],
    ["Branch", "Lihat semua branch", "git branch -a", "Melihat semua branch (local dan remote)"],
    ["Branch", "Hapus branch lokal", "git branch -d nama-branch", "Menghapus branch yang sudah di-merge"],
    ["Branch", "Hapus branch remote", "git push origin --delete nama-branch", "Menghapus branch di remote"],
    ["Status", "Cek status", "git status", "Melihat status working directory"],
    ["Log", "Lihat history commit", "git log --oneline -10", "Melihat 10 commit terakhir"],
    ["Diff", "Lihat perubahan", "git diff main...nama-branch", "Melihat perbedaan dengan main"],
]

for row_idx, row_data in enumerate(data5, 4):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws5.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border

auto_width(ws5)

# ============ SHEET 6: Template Komentar ============
ws6 = wb.create_sheet("6. Template Komentar")

ws6['A1'] = "TEMPLATE KOMENTAR REVIEW"
ws6['A1'].font = title_font
ws6['A2'] = "Copy-paste template saat memberikan review di GitHub"

headers6 = ["Tipe", "Tag", "Template Komentar", "Kapan Digunakan"]
for col, header in enumerate(headers6, 1):
    ws6.cell(row=4, column=col, value=header)
apply_header_style(ws6, 4, len(headers6))

data6 = [
    ["Approve", "[LGTM]", "LGTM! Code sudah bagus dan siap merge. ✅", "PR siap merge"],
    ["Minor Improvement", "[NIT]", "[NIT] Bisa diperbaiki tapi tidak blocking: [detail]", "Improvement kecil"],
    ["Request Change", "[BLOCKER]", "[BLOCKER] Perlu diperbaiki sebelum merge: [detail]", "HARUS fix dulu"],
    ["Security Issue", "[SECURITY]", "[SECURITY] ⚠️ Ada risiko keamanan: [detail]", "Security issue"],
    ["Bug Found", "[BUG]", "[BUG] 🐛 Sepertinya ada bug: [detail]", "Bug ditemukan"],
    ["Question", "[Q]", "[Q] Bisa dijelaskan kenapa menggunakan approach ini?", "Butuh klarifikasi"],
    ["Suggestion", "[SUGGESTION]", "[SUGGESTION] Pertimbangkan menggunakan [alternatif]", "Ada cara lebih baik"],
    ["Missing Test", "[TEST]", "[TEST] Mohon tambahkan unit test untuk logic ini", "Test tidak ada"],
    ["Praise", "[NICE]", "[NICE] 👍 Approach yang bagus! Clean dan readable.", "Code bagus"],
]

for row_idx, row_data in enumerate(data6, 5):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws6.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border

auto_width(ws6)

# ============ SHEET 7: Escalation Matrix ============
ws7 = wb.create_sheet("7. Escalation Matrix")

ws7['A1'] = "ESCALATION MATRIX"
ws7['A1'].font = title_font

headers7 = ["Kondisi", "Aksi", "Eskalasi Ke", "SLA", "Template Komunikasi"]
for col, header in enumerate(headers7, 1):
    ws7.cell(row=3, column=col, value=header)
apply_header_style(ws7, 3, len(headers7))

data7 = [
    ["PR 3+ hari tidak di-review", "Kirim reminder", "Lead/Manager", "1 hari", "Mohon review PR #[nomor]"],
    ["Conflict kompleks", "Minta bantuan resolve", "Author PR", "2 hari", "Ada conflict, bisa bantu resolve?"],
    ["Security issue", "Block merge SEGERA", "Security Lead", "< 4 jam", "[URGENT] Security issue di PR #[nomor]"],
    ["Tidak yakin dengan logic", "Diskusi dengan author", "Author/Lead", "1 hari", "Bisa jelaskan logic di bagian [X]?"],
    ["CI/CD gagal berulang", "Investigasi bersama", "DevOps/Author", "1 hari", "CI/CD failing, butuh bantuan"],
    ["PR terlalu besar (500+ lines)", "Minta split PR", "Author PR", "Sebelum review", "Bisa split jadi PR lebih kecil?"],
]

for row_idx, row_data in enumerate(data7, 4):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws7.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border

auto_width(ws7)

# Save workbook
output_path = "docs/SOP_PR_Review_Workbook.xlsx"
wb.save(output_path)
print(f"✅ Workbook berhasil dibuat: {output_path}")
print(f"   - 7 sheets dalam 1 file")
print(f"   - Bisa langsung dibuka dengan Excel/Google Sheets")
